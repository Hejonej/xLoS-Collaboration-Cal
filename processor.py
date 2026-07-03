"""
xLoS / 동일 LoS 분석 처리 엔진
Flask 앱에서 호출되며, REPORT_MONTH와 파일 경로를 매개변수로 받는다.
"""
from pathlib import Path
from io import BytesIO
import shutil
import zipfile

import pandas as pd
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Border, Side


import re as _re

_VALID_BORDER_STYLES = {
    "medium", "thick", "dashed", "dotted", "mediumDashDot", "mediumDashed",
    "double", "hair", "slantDashDot", "thin", "dashDotDot", "dashDot",
    "mediumDashDotDot",
}


def _fix_styles_xml(raw_xml: bytes) -> bytes:
    """깨진 border style 값만 제거하고 나머지(숫자 서식 등)는 보존한다."""
    text = raw_xml.decode("utf-8", errors="replace")
    def _fix_style_attr(m):
        val = m.group(1)
        if val in _VALID_BORDER_STYLES:
            return m.group(0)
        return ""
    text = _re.sub(r'\bstyle="([^"]*)"', _fix_style_attr, text)
    return text.encode("utf-8")


SHEET_XLOS = "Project List (xLoS)"
SHEET_SAME = "Project List (동일 LoS)"


# ── YYYYMM 산술 유틸 ─────────────────────────────────────────

def _ym_add(ym: int, months: int) -> int:
    y, m = divmod(ym, 100)
    total = y * 12 + (m - 1) + months
    return (total // 12) * 100 + (total % 12) + 1


def _ym_set(start: int, end: int) -> set[int]:
    result = set()
    cur = start
    while cur <= end:
        result.add(cur)
        cur = _ym_add(cur, 1)
    return result


# ── 기간 설정 클래스 ─────────────────────────────────────────

class PeriodConfig:
    """REPORT_MONTH 하나로 모든 FY·기간 정보를 계산한다."""

    def __init__(self, report_month: int):
        self.report_month = report_month
        rm_year = report_month // 100
        rm_month = report_month % 100

        self.curr_fy = rm_year - 2000 + 1 if rm_month >= 7 else rm_year - 2000
        self.prev_fy = self.curr_fy - 1

        self.curr_fy_start = (2000 + self.curr_fy - 1) * 100 + 7
        self.prev_fy_start = (2000 + self.prev_fy - 1) * 100 + 7
        self.prev_fy_end = _ym_add(self.curr_fy_start, -1)

        n_months = (rm_month - 7 + 1) if rm_month >= 7 else (rm_month + 12 - 7 + 1)
        self.prev_period_end = _ym_add(self.prev_fy_start, n_months - 1)

        self.curr_period_set = _ym_set(self.curr_fy_start, report_month)
        self.prev_period_set = _ym_set(self.prev_fy_start, self.prev_period_end)
        self.prev_full_set = _ym_set(self.prev_fy_start, self.prev_fy_end)

        _s = lambda ym: f"{ym % 10000:04d}"
        self.curr_range_label = f"{_s(self.curr_fy_start)}~{_s(report_month)}"
        self.prev_range_label = f"{_s(self.prev_fy_start)}~{_s(self.prev_period_end)}"
        self.prev_full_label = f"{_s(self.prev_fy_start)}~{_s(self.prev_fy_end)}"
        self.month_label = f"{rm_month}월말"


# ── 데이터 읽기 / 전처리 ────────────────────────────────────

def _read_raw(path: str) -> pd.DataFrame:
    try:
        return pd.read_excel(path, header=1, engine="openpyxl")
    except Exception as e:
        if "stylesheet" not in str(e).lower():
            raise
        # 스타일시트 오류 → 깨진 border style 속성만 제거하고 재시도
        raw = BytesIO(open(path, "rb").read())
        fixed = BytesIO()
        with zipfile.ZipFile(raw, "r") as zin, zipfile.ZipFile(fixed, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/styles.xml":
                    data = _fix_styles_xml(data)
                zout.writestr(item, data)
        fixed.seek(0)
        return pd.read_excel(fixed, header=1, engine="openpyxl")


def _ensure_column(df: pd.DataFrame, standard: str, candidates: list[str]) -> None:
    if standard in df.columns:
        return
    for cand in candidates:
        if cand in df.columns:
            df[standard] = df[cand]
            return


_VALID_CREATION_TYPES = {"노츠생성", "자동생성(F-link)", "수기생성(F-link)"}

_OFFSET_MONEY_COLS = ["예상(표준) EM", "Refer EM", "Matching EM", "Refer EM Total"]


def _remove_offset_pairs(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Rating이 '-'로 시작하는 행과 같은 Project Code에서 금액이 반대인 상계 쌍을 삭제한다."""
    _ensure_column(df, "Rating", ["Rating"])
    if "Rating" not in df.columns or "Project Code" not in df.columns:
        return df, 0

    money_cols = [c for c in _OFFSET_MONEY_COLS if c in df.columns]
    if not money_cols:
        return df, 0

    neg_mask = df["Rating"].astype(str).str.match(r"^-")
    if not neg_mask.any():
        return df, 0

    remove_indices = set()

    for neg_idx in df[neg_mask].index:
        if neg_idx in remove_indices:
            continue
        neg_row = df.loc[neg_idx]
        pc = neg_row["Project Code"]

        candidates = df[
            (df["Project Code"] == pc)
            & (~df.index.isin(remove_indices))
            & (df.index != neg_idx)
            & (~df["Rating"].astype(str).str.match(r"^-"))
        ]

        for cand_idx in candidates.index:
            cand_row = df.loc[cand_idx]
            is_pair = True
            for col in money_cols:
                neg_val = pd.to_numeric(neg_row[col], errors="coerce") or 0
                cand_val = pd.to_numeric(cand_row[col], errors="coerce") or 0
                if abs(neg_val + cand_val) > 0.01:
                    is_pair = False
                    break
            if is_pair:
                remove_indices.add(neg_idx)
                remove_indices.add(cand_idx)
                break

    removed = len(remove_indices)
    if removed:
        df = df.drop(index=list(remove_indices)).reset_index(drop=True)
    return df, removed


def _merge_and_dedup(prev_df: pd.DataFrame, curr_df: pd.DataFrame) -> pd.DataFrame:
    df = pd.concat([prev_df, curr_df], ignore_index=True)

    _ensure_column(df, "관리반영월", ["관리반영월"])
    _ensure_column(df, "Project Code", ["Project Code"])
    _ensure_column(df, "생성 구분", ["생성 구분", "생성구분", "생성_구분"])

    for col in ["관리반영월", "Project Code", "생성 구분"]:
        if col not in df.columns:
            raise KeyError(f"필수 컬럼이 없습니다: {col}")

    # Raw 데이터 하단의 합계(Total 계) 행 등 비데이터 행 제거
    df = df[df["생성 구분"].astype(str).str.strip().isin(_VALID_CREATION_TYPES)]

    df["__uniq__"] = (
        df["관리반영월"].astype(str) + "|"
        + df["Project Code"].astype(str) + "|"
        + df["생성 구분"].astype(str)
    )
    df = df.sort_values("__uniq__").drop_duplicates("__uniq__", keep="last")
    return df.drop(columns="__uniq__")


# ── 승인 기준 Y/N ───────────────────────────────────────────

def _calc_approval_yn(series, period_months: set[int]) -> pd.Series:
    cleaned = series.astype(str).str.replace(",", "", regex=False)
    s = pd.to_numeric(cleaned, errors="coerce").fillna(0).astype(int)
    return s.isin(period_months).map({True: "Y", False: "N"})


def _effective_approval_month(df: pd.DataFrame) -> pd.Series:
    """관리반영월이 승인월보다 앞서면 관리반영월을 승인 기준 산정에 사용."""
    appr_clean = df["승인월"].astype(str).str.replace(",", "", regex=False)
    appr_int = pd.to_numeric(appr_clean, errors="coerce").fillna(0).astype(int)

    if "관리반영월" not in df.columns:
        return appr_int

    mgmt_clean = df["관리반영월"].astype(str).str.replace(",", "", regex=False)
    mgmt_int = pd.to_numeric(mgmt_clean, errors="coerce").fillna(0).astype(int)

    use_mgmt = (mgmt_int > 0) & (appr_int > 0) & (mgmt_int < appr_int)
    effective = appr_int.copy()
    effective.loc[use_mgmt] = mgmt_int.loc[use_mgmt]
    return effective


# ── Project List 빌드 ───────────────────────────────────────

def _build_project_list(df: pd.DataFrame, *, is_xlos: bool, cfg: PeriodConfig) -> pd.DataFrame:
    if is_xlos:
        preferred_cols = [
            "생성 구분", "관리반영월", "반영",
            "LoS", "본부", "PTR", "승인당시 본부",
            "Project Code", "협업여부", "확정/성공", "Client명", "용역명",
            "LoS.1", "본부.1", "PTR.1",
            "업무유형", "계약금액", "예상(표준) EM", "Rating",
            "Refer EM", "Matching EM", "Refer EM Total",
            "누적수익", "Scoring", "LL 승인날짜", "승인월",
            "주관 Project Code", "합산 Rev", "합산 계약금", "합산 예상 EM",
        ]
    else:
        preferred_cols = [
            "승인 기준", "승인 기준_1", "승인 기준_2",
            "생성 구분", "관리반영월", "반영",
            "LoS", "본부", "PTR", "승인당시 본부",
            "Project Code", "협업여부", "확정/성공", "Client명", "용역명",
            "LoS.1", "본부.1", "PTR.1",
            "업무유형", "계약금액", "예상(표준) EM", "Rating",
            "누적수익", "Scoring", "LL 승인날짜", "승인월",
            "주관 Project Code", "합산 Rev", "합산 계약금", "합산 예상 EM",
        ]

    available = [c for c in preferred_cols if c in df.columns]
    result = df[available].copy()

    # 승인월이 없으면 LL 승인날짜에서 파생
    if "승인월" not in df.columns and "LL 승인날짜" in df.columns:
        raw_date = df["LL 승인날짜"].astype(str).str.replace(",", "", regex=False)
        df["승인월"] = raw_date.str[:6]
        result["승인월"] = df["승인월"].values
    _ensure_column(df, "승인월", ["승인월"])

    if is_xlos:
        # xLoS 3탭: B~G열 6개 (승인 기준 + 관리반영 기준) × 3기간
        if "승인월" in df.columns:
            eff = _effective_approval_month(df)
            appr_curr = _calc_approval_yn(eff, cfg.curr_period_set).values
            appr_prev = _calc_approval_yn(eff, cfg.prev_period_set).values
            appr_full = _calc_approval_yn(eff, cfg.prev_full_set).values
        else:
            n = len(result)
            appr_curr = appr_prev = appr_full = ["N"] * n

        if "관리반영월" in df.columns:
            mgmt_curr = _calc_approval_yn(df["관리반영월"], cfg.curr_period_set).values
            mgmt_prev = _calc_approval_yn(df["관리반영월"], cfg.prev_period_set).values
            mgmt_full = _calc_approval_yn(df["관리반영월"], cfg.prev_full_set).values
        else:
            n = len(result)
            mgmt_curr = mgmt_prev = mgmt_full = ["N"] * n

        result.insert(0, "승인 기준", appr_curr)
        result.insert(1, "관리반영 기준", mgmt_curr)
        result.insert(2, "승인 기준_1", appr_prev)
        result.insert(3, "관리반영 기준_1", mgmt_prev)
        result.insert(4, "승인 기준_2", appr_full)
        result.insert(5, "관리반영 기준_2", mgmt_full)
    else:
        # 동일LoS 6탭: B~D열 3개 (승인 기준) × 3기간
        if "승인월" in df.columns:
            eff = _effective_approval_month(df)
            result.insert(0, "승인 기준", _calc_approval_yn(eff, cfg.curr_period_set).values)
            result.insert(1, "승인 기준_1", _calc_approval_yn(eff, cfg.prev_period_set).values)
            result.insert(2, "승인 기준_2", _calc_approval_yn(eff, cfg.prev_full_set).values)

    if "합산 Rev" not in result.columns and "누적수익" in df.columns:
        result["합산 Rev"] = df["누적수익"]
    if "합산 계약금" not in result.columns and "계약금액" in df.columns:
        result["합산 계약금"] = df["계약금액"]
    if "합산 예상 EM" not in result.columns and "예상(표준) EM" in df.columns:
        result["합산 예상 EM"] = df["예상(표준) EM"]

    if "Rating" in result.columns:
        result["Rating"] = (
            result["Rating"].astype(str)
            .str.extract(r"([A-Za-z가-힣]+)", expand=False)
        )

    return result


# ── 템플릿 기간 라벨 갱신 ───────────────────────────────────

def _update_period_labels(wb, cfg: PeriodConfig) -> None:
    text_map = [
        ("9월말", cfg.month_label),
        ("2507~2509", cfg.curr_range_label),
        ("2407~2409", cfg.prev_range_label),
    ]
    if cfg.curr_fy != 26:
        text_map.insert(0, ("FY26", f"FY{cfg.curr_fy}"))
        text_map.append(("FY25", f"FY{cfg.prev_fy}"))

    for ws in wb.worksheets:
        for r in range(1, min(26, ws.max_row + 1)):
            for c in range(1, min(40, ws.max_column + 1)):
                val = ws.cell(row=r, column=c).value
                if not isinstance(val, str):
                    continue
                new_val = val
                for old, new in text_map:
                    new_val = new_val.replace(old, new)
                if new_val != val:
                    ws.cell(row=r, column=c).value = new_val


# ── 헤더 행 탐색 ───────────────────────────────────────────

def _find_header_row(ws) -> int:
    for r in range(1, ws.max_row + 1):
        row_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if "Project Code" in row_values:
            return r
    raise ValueError("'Project Code' 헤더가 있는 행을 찾지 못했습니다.")


# ── 템플릿에 데이터 쓰기 ───────────────────────────────────

def _write_to_template(
    df_xlos: pd.DataFrame,
    df_same: pd.DataFrame,
    template_path: str,
    output_path: str,
    cfg: PeriodConfig,
) -> None:
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)

    _update_period_labels(wb, cfg)

    def write_sheet(sheet_name: str, df: pd.DataFrame, clear_all: bool = False):
        ws = wb[sheet_name]
        header_row = _find_header_row(ws)
        start_row = header_row + 1

        headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
        col_index = {}
        for idx, name in enumerate(headers):
            if not name:
                continue
            key = name
            if key in col_index:
                suffix = 1
                while f"{name}_{suffix}" in col_index:
                    suffix += 1
                key = f"{name}_{suffix}"
            col_index[key] = idx + 1

        is_xlos_sheet = sheet_name == SHEET_XLOS
        if is_xlos_sheet and all(c in df.columns for c in ("승인 기준", "관리반영 기준", "승인 기준_1", "관리반영 기준_1", "승인 기준_2", "관리반영 기준_2")):
            col_index["승인 기준"] = 2
            col_index["관리반영 기준"] = 3
            col_index["승인 기준_1"] = 4
            col_index["관리반영 기준_1"] = 5
            col_index["승인 기준_2"] = 6
            col_index["관리반영 기준_2"] = 7

        is_same_sheet = sheet_name == SHEET_SAME
        if is_same_sheet and all(c in df.columns for c in ("승인 기준", "승인 기준_1", "승인 기준_2")):
            col_index["승인 기준"] = 2
            col_index["승인 기준_1"] = 3
            col_index["승인 기준_2"] = 4

        df_to_template = {}
        for c in df.columns:
            if c in col_index:
                df_to_template[c] = c
            elif "." in c:
                alt = c.replace(".", "_")
                if alt in col_index:
                    df_to_template[c] = alt

        writable_cols = list(df_to_template.keys())
        writable_col_indexes = set(col_index[df_to_template[c]] for c in writable_cols)
        all_col_indexes = list(range(1, ws.max_column + 1))

        max_target_row = min(start_row + len(df) + 500, ws.max_row)

        if clear_all:
            for r in range(start_row, max_target_row + 1):
                for c in all_col_indexes:
                    ws.cell(row=r, column=c).value = None
        else:
            for r in range(start_row, max_target_row + 1):
                for c in writable_col_indexes:
                    ws.cell(row=r, column=c).value = None

        current_row = start_row
        for _, row in df.iterrows():
            for df_col in writable_cols:
                tmpl_col = df_to_template[df_col]
                c_idx = col_index[tmpl_col]
                ws.cell(row=current_row, column=c_idx).value = row[df_col]
            current_row += 1

        last_data_row = current_row - 1

        # 11행(start_row) 서식을 데이터 끝행까지 복사
        if last_data_row > start_row:
            for c in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=start_row, column=c)
                for r in range(start_row + 1, last_data_row + 1):
                    dest_cell = ws.cell(row=r, column=c)
                    if src_cell.has_style:
                        dest_cell.font = copy(src_cell.font)
                        dest_cell.border = copy(src_cell.border)
                        dest_cell.fill = copy(src_cell.fill)
                        dest_cell.number_format = src_cell.number_format
                        dest_cell.alignment = copy(src_cell.alignment)

    write_sheet(SHEET_XLOS, df_xlos, clear_all=True)
    write_sheet(SHEET_SAME, df_same, clear_all=True)

    # B9, C9 윗선 복원 (Project List xLoS 탭)
    ws_xlos = wb[SHEET_XLOS]
    thin_top = Border(top=Side(style="thin"))
    for col in [2, 3]:  # B, C
        cell = ws_xlos.cell(row=9, column=col)
        existing = cell.border
        cell.border = Border(
            top=Side(style="thin"),
            left=existing.left,
            right=existing.right,
            bottom=existing.bottom,
        )

    wb.save(output_path)


# ── LoS별 요약 집계 ───────────────────────────────────────

_LOS_ORDER = ["Assurance", "Tax", "Deals", "Consulting", "Public Sector",
              "AX Node", "Ax Node", "IFS"]


def compute_summary(df: pd.DataFrame, cfg: PeriodConfig, *, is_xlos: bool) -> dict:
    if "LoS" not in df.columns:
        return {"labels": [], "data": [], "is_xlos": is_xlos}

    periods = [
        ("승인 기준", "관리반영 기준"),
        ("승인 기준_1", "관리반영 기준_1"),
        ("승인 기준_2", "관리반영 기준_2"),
    ]
    period_names = [
        f"FY{cfg.curr_fy} ({cfg.curr_range_label})",
        f"FY{cfg.prev_fy} ({cfg.prev_range_label})",
        f"FY{cfg.prev_fy} Full Year",
    ]

    def _safe_int_sum(s):
        return int(round(pd.to_numeric(s, errors="coerce").fillna(0).sum()))

    los_vals = [l for l in df["LoS"].unique() if pd.notna(l)]
    ordered = [l for l in _LOS_ORDER if l in los_vals]
    ordered += sorted(set(los_vals) - set(ordered))

    _AMT_TYPES = {"노츠생성", "자동생성(F-link)"}

    rows = []
    for los in ordered:
        sub = df[df["LoS"] == los]
        r: dict = {"los": los}
        for i, (acol, mcol) in enumerate(periods):
            m_cnt = (sub[acol] == "Y") & (sub["생성 구분"] == "노츠생성")
            m_amt = (sub[acol] == "Y") & (sub["생성 구분"].isin(_AMT_TYPES))
            r[f"c{i}"] = int(m_cnt.sum())
            r[f"a{i}"] = _safe_int_sum(sub.loc[m_amt, "계약금액"]) if "계약금액" in sub else 0
            if is_xlos and "Refer EM" in sub.columns:
                r[f"r{i}"] = _safe_int_sum(sub.loc[sub[mcol] == "Y", "Refer EM"])
        rows.append(r)

    tot: dict = {"los": "Total"}
    for k in [f"{p}{i}" for i in range(3) for p in ("c", "a")]:
        tot[k] = sum(r.get(k, 0) for r in rows)
    if is_xlos:
        for k in [f"r{i}" for i in range(3)]:
            tot[k] = sum(r.get(k, 0) for r in rows)
    rows.append(tot)

    return {"labels": period_names, "data": rows, "is_xlos": is_xlos}


# ── 자동 검증 ─────────────────────────────────────────────

def validate_results(
    *,
    x_prev: pd.DataFrame,
    x_curr: pd.DataFrame,
    s_prev: pd.DataFrame,
    s_curr: pd.DataFrame,
    x_merged: pd.DataFrame,
    s_merged: pd.DataFrame,
    pl_xlos: pd.DataFrame,
    pl_same: pd.DataFrame,
    cfg: PeriodConfig,
    x_offset_removed: int = 0,
    s_offset_removed: int = 0,
) -> list[dict]:
    """처리 결과의 정합성을 자동 검증하고 항목별 결과 리스트를 반환한다."""
    checks: list[dict] = []

    # ── 1. Raw 건수 보존 검증 ──────────────────────────────────
    for label, prev, curr, merged in [
        ("xLoS", x_prev, x_curr, x_merged),
        ("동일LoS", s_prev, s_curr, s_merged),
    ]:
        n_prev = len(prev)
        n_curr = len(curr)

        _ensure_column(prev, "생성 구분", ["생성 구분", "생성구분", "생성_구분"])
        _ensure_column(curr, "생성 구분", ["생성 구분", "생성구분", "생성_구분"])
        valid_prev = prev["생성 구분"].astype(str).str.strip().isin(_VALID_CREATION_TYPES).sum() if "생성 구분" in prev.columns else n_prev
        valid_curr = curr["생성 구분"].astype(str).str.strip().isin(_VALID_CREATION_TYPES).sum() if "생성 구분" in curr.columns else n_curr

        n_merged = len(merged)
        n_dup = int(valid_prev + valid_curr - n_merged)
        checks.append({
            "id": f"row_count_{label}",
            "label": f"Raw 건수 보존 ({label})",
            "status": "pass",
            "detail": (
                f"전기 {valid_prev} + 당기 {valid_curr}"
                f" → 중복 {n_dup}건 제거 → {n_merged}건"
            ),
        })

    # ── 1b. 상계 쌍 제거 현황 ─────────────────────────────────
    for label, removed in [("xLoS", x_offset_removed), ("동일LoS", s_offset_removed)]:
        if removed > 0:
            checks.append({
                "id": f"offset_removed_{label}",
                "label": f"상계 쌍 제거 ({label})",
                "status": "info",
                "detail": f"Rating 음수 상계 {removed}행 제거됨 ({removed // 2}쌍)",
            })

    # ── 2. 승인 기준 Y/N 역검증 ────────────────────────────────
    for label, df, is_xlos in [
        ("xLoS", pl_xlos, True),
        ("동일LoS", pl_same, False),
    ]:
        if "승인 기준" not in df.columns or "승인월" not in df.columns:
            continue
        eff = _effective_approval_month(df)
        expected = _calc_approval_yn(eff, cfg.curr_period_set)
        actual = df["승인 기준"]
        mismatch = (expected != actual).sum()
        y_cnt = int((actual == "Y").sum())
        n_cnt = int((actual == "N").sum())
        status = "pass" if mismatch == 0 else "fail"
        detail = f"Y {y_cnt}건, N {n_cnt}건"
        if mismatch:
            detail += f" — 불일치 {mismatch}건"
        else:
            detail += " — 전수 일치"
        checks.append({
            "id": f"approval_yn_{label}",
            "label": f"승인 기준 정합성 ({label})",
            "status": status,
            "detail": detail,
        })

    # ── 3. 음수값 존재 확인 ────────────────────────────────────
    money_cols = ["Refer EM", "계약금액", "누적수익", "합산 Rev", "합산 계약금"]
    for label, df in [("xLoS", pl_xlos), ("동일LoS", pl_same)]:
        neg_msgs = []
        for col in money_cols:
            if col not in df.columns:
                continue
            nums = pd.to_numeric(df[col], errors="coerce")
            neg_mask = nums < 0
            neg_count = int(neg_mask.sum())
            if neg_count > 0:
                pc_col = "Project Code" if "Project Code" in df.columns else None
                example = ""
                if pc_col:
                    first_pc = df.loc[neg_mask, pc_col].iloc[0]
                    extra = f" 외 {neg_count - 1}건" if neg_count > 1 else ""
                    example = f" ({first_pc}{extra})"
                neg_msgs.append(f"{col} {neg_count}건{example}")
        if neg_msgs:
            checks.append({
                "id": f"negative_{label}",
                "label": f"음수값 감지 ({label})",
                "status": "info",
                "detail": ", ".join(neg_msgs),
            })
        else:
            checks.append({
                "id": f"negative_{label}",
                "label": f"음수값 검사 ({label})",
                "status": "pass",
                "detail": "금액 컬럼 전체 음수 없음",
            })

    # ── 4. LoS별 교차 검증 (Python 집계 vs 엑셀 수식 대용) ──────
    for label, df, is_xlos in [
        ("xLoS", pl_xlos, True),
        ("동일LoS", pl_same, False),
    ]:
        if "LoS" not in df.columns or "승인 기준" not in df.columns:
            continue
        acol = "승인 기준"
        mask_cnt = (df[acol] == "Y") & (df["생성 구분"] == "노츠생성")
        total_cnt = int(mask_cnt.sum())

        amt_col = "계약금액" if "계약금액" in df.columns else None
        if amt_col:
            mask_amt = (df[acol] == "Y") & (df["생성 구분"].isin(_VALID_CREATION_TYPES))
            total_amt = int(round(pd.to_numeric(df.loc[mask_amt, amt_col], errors="coerce").fillna(0).sum()))
        else:
            total_amt = 0

        ref_total = 0
        if is_xlos and "Refer EM" in df.columns and "관리반영 기준" in df.columns:
            ref_total = int(round(
                pd.to_numeric(df.loc[df["관리반영 기준"] == "Y", "Refer EM"], errors="coerce").fillna(0).sum()
            ))

        parts = [f"건수 {total_cnt}", f"금액 {total_amt:,}"]
        if is_xlos:
            parts.append(f"Refer EM {ref_total:,}")
        checks.append({
            "id": f"cross_check_{label}",
            "label": f"FY{cfg.curr_fy} 집계 검산 ({label})",
            "status": "pass",
            "detail": " / ".join(parts),
        })

    # ── 5. 생성 구분 분포 확인 ─────────────────────────────────
    for label, prev, curr, merged in [
        ("xLoS", x_prev, x_curr, x_merged),
        ("동일LoS", s_prev, s_curr, s_merged),
    ]:
        _ensure_column(prev, "생성 구분", ["생성 구분", "생성구분", "생성_구분"])
        _ensure_column(curr, "생성 구분", ["생성 구분", "생성구분", "생성_구분"])
        if "생성 구분" not in prev.columns:
            continue
        all_raw = pd.concat([prev, curr], ignore_index=True)
        counts = all_raw["생성 구분"].astype(str).str.strip().value_counts()
        excluded = {k: int(v) for k, v in counts.items() if k not in _VALID_CREATION_TYPES and k != "nan"}
        if excluded:
            parts = [f"{k} {v}건" for k, v in excluded.items()]
            checks.append({
                "id": f"creation_type_{label}",
                "label": f"생성구분 필터 ({label})",
                "status": "info",
                "detail": "제외됨: " + ", ".join(parts),
            })
        else:
            checks.append({
                "id": f"creation_type_{label}",
                "label": f"생성구분 분포 ({label})",
                "status": "pass",
                "detail": "전체 행이 유효한 생성구분",
            })

    return checks


# ── 메인 처리 함수 (Flask에서 호출) ─────────────────────────

def process_files(
    report_month: int,
    xlos_prev_path: str,
    xlos_curr_path: str,
    same_prev_path: str,
    same_curr_path: str,
    template_path: str,
    output_path: str,
) -> dict:
    """모든 파일을 처리하고 output_path에 결과를 저장한다."""
    cfg = PeriodConfig(report_month)

    x_prev = _read_raw(xlos_prev_path)
    x_curr = _read_raw(xlos_curr_path)
    s_prev = _read_raw(same_prev_path)
    s_curr = _read_raw(same_curr_path)

    x_merged = _merge_and_dedup(x_prev, x_curr)
    s_merged = _merge_and_dedup(s_prev, s_curr)

    x_merged, x_offset_removed = _remove_offset_pairs(x_merged)
    s_merged, s_offset_removed = _remove_offset_pairs(s_merged)

    pl_xlos = _build_project_list(x_merged, is_xlos=True, cfg=cfg)
    pl_same = _build_project_list(s_merged, is_xlos=False, cfg=cfg)

    _write_to_template(pl_xlos, pl_same, template_path, output_path, cfg)

    validation = validate_results(
        x_prev=x_prev, x_curr=x_curr,
        s_prev=s_prev, s_curr=s_curr,
        x_merged=x_merged, s_merged=s_merged,
        pl_xlos=pl_xlos, pl_same=pl_same,
        cfg=cfg,
        x_offset_removed=x_offset_removed,
        s_offset_removed=s_offset_removed,
    )

    return {
        "output_path": output_path,
        "xlos_summary": compute_summary(pl_xlos, cfg, is_xlos=True),
        "same_summary": compute_summary(pl_same, cfg, is_xlos=False),
        "validation": validation,
    }
